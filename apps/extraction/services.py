import re
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pandas as pd
import pdfplumber
import pytesseract
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image

from apps.dashboard.services import log_activity
from apps.extraction.models import ExtractionResult
from apps.subscriptions.services import enforce_daily_number_limit
from apps.uploads.models import UploadedFile

PHONE_PATTERN = re.compile(r"(\+?\d[\d \t-]{8,15}\d)")
CSV_CHUNK_SIZE = 500


def iter_pdf_text(uploaded_file):
    with pdfplumber.open(uploaded_file.file.path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text:
                yield text


def iter_xlsx_text(uploaded_file):
    workbook = load_workbook(uploaded_file.file.path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    yield " ".join(values)
    finally:
        workbook.close()


def iter_excel_text(uploaded_file):
    extension = Path(uploaded_file.file.name).suffix.lower()
    if extension == ".xlsx":
        yield from iter_xlsx_text(uploaded_file)
        return

    workbook = pd.ExcelFile(uploaded_file.file.path)
    for sheet_name in workbook.sheet_names:
        frame = workbook.parse(sheet_name=sheet_name, dtype=str).fillna("")
        if not frame.empty or list(frame.columns):
            yield frame.to_string(index=False)


def iter_csv_text(uploaded_file):
    for frame in pd.read_csv(uploaded_file.file.path, dtype=str, chunksize=CSV_CHUNK_SIZE):
        frame = frame.fillna("")
        if not frame.empty or list(frame.columns):
            yield frame.to_string(index=False)


def iter_image_text(uploaded_file):
    with Image.open(uploaded_file.file.path) as image:
        text = pytesseract.image_to_string(image)
        if text:
            yield text


def iter_uploaded_content(uploaded_file):
    readers = {
        UploadedFile.FileTypeChoices.PDF: iter_pdf_text,
        UploadedFile.FileTypeChoices.EXCEL: iter_excel_text,
        UploadedFile.FileTypeChoices.CSV: iter_csv_text,
        UploadedFile.FileTypeChoices.IMAGE: iter_image_text,
    }
    yield from readers[uploaded_file.file_type](uploaded_file)


def normalize_phone_number(raw_number):
    digits = re.sub(r"\D", "", raw_number)
    if not digits:
        return None

    if digits.startswith("00"):
        digits = digits[2:]

    if digits.startswith("971") and len(digits) == 12:
        return f"+971{digits[3:]}"
    if digits.startswith("91") and len(digits) == 12:
        return f"+91{digits[2:]}"
    if len(digits) == 10 and digits.startswith("05"):
        return f"+971{digits[1:]}"
    if len(digits) == 9 and digits.startswith("5"):
        return f"+971{digits}"
    if len(digits) == 11 and digits.startswith("0") and digits[1] == "5":
        return f"+971{digits[1:]}"
    if len(digits) == 11 and digits.startswith("0"):
        return f"+91{digits[1:]}"
    if len(digits) == 10:
        return f"+91{digits}"
    return None


def extract_numbers(raw_text):
    return extract_numbers_from_chunks([raw_text])


def extract_numbers_from_chunks(chunks):
    extracted_numbers = []
    seen_numbers = set()

    for chunk in chunks:
        if not chunk:
            continue
        for match in PHONE_PATTERN.findall(chunk):
            normalized_number = normalize_phone_number(match)
            if normalized_number and normalized_number not in seen_numbers:
                seen_numbers.add(normalized_number)
                extracted_numbers.append(normalized_number)

    return extracted_numbers


def extract_numbers_from_upload(uploaded_file):
    return extract_numbers_from_chunks(iter_uploaded_content(uploaded_file))


def export_numbers_to_excel(result):
    dataframe = pd.DataFrame({"phone_number": result.numbers})
    buffer = BytesIO()
    dataframe.to_excel(buffer, index=False)
    buffer.seek(0)

    result_filename = f"{Path(result.upload.original_name).stem[:50]}_{uuid4().hex}.xlsx"
    if result.result_file:
        result.result_file.delete(save=False)
    result.result_file.save(result_filename, ContentFile(buffer.getvalue()), save=False)
    result.save(update_fields=["result_file", "updated_at"])
    return result


def process_uploaded_file(uploaded_file):
    uploaded_file.status = UploadedFile.StatusChoices.PROCESSING
    uploaded_file.processed_at = timezone.now()
    uploaded_file.error_message = ""
    uploaded_file.save(update_fields=["status", "processed_at", "error_message"])

    try:
        numbers = extract_numbers_from_upload(uploaded_file)
        try:
            existing_total = uploaded_file.extraction_result.total_numbers
        except UploadedFile.extraction_result.RelatedObjectDoesNotExist:
            existing_total = 0
        enforce_daily_number_limit(uploaded_file.user, len(numbers), existing_total=existing_total)

        result, _ = ExtractionResult.objects.update_or_create(
            upload=uploaded_file,
            defaults={
                "user": uploaded_file.user,
                "numbers": numbers,
                "total_numbers": len(numbers),
            },
        )
        export_numbers_to_excel(result)

        uploaded_file.status = UploadedFile.StatusChoices.COMPLETED
        uploaded_file.processed_at = timezone.now()
        uploaded_file.error_message = ""
        uploaded_file.save(update_fields=["status", "processed_at", "error_message"])

        log_activity(
            user=uploaded_file.user,
            action="extraction_completed",
            description=f"Processed {uploaded_file.original_name}.",
            metadata={
                "upload_id": str(uploaded_file.id),
                "result_id": str(result.id),
                "total_numbers": result.total_numbers,
            },
        )
        return result
    except Exception as exc:
        uploaded_file.status = UploadedFile.StatusChoices.FAILED
        uploaded_file.error_message = str(exc)
        uploaded_file.processed_at = timezone.now()
        uploaded_file.save(update_fields=["status", "error_message", "processed_at"])

        log_activity(
            user=uploaded_file.user,
            action="extraction_failed",
            description=f"Failed to process {uploaded_file.original_name}.",
            metadata={"upload_id": str(uploaded_file.id), "error": str(exc)},
        )
        raise
